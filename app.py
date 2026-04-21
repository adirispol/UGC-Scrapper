import streamlit as st
import pandas as pd
import requests
import time
import io
import html
import os
from datetime import datetime, timedelta, timezone, date
import urllib.parse
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CONFIG
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ACTOR_LINKEDIN = "supreme_coder~linkedin-post"
ACTOR_X = "xquik~x-tweet-scraper"
APIFY_BASE = "https://api.apify.com/v2"
IST = timezone(timedelta(hours=5, minutes=30))

COST_PER_LINKEDIN_POST = 0.005
COST_PER_X_TWEET = 0.00015

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PAGE CONFIG — must be FIRST streamlit call
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
st.set_page_config(
    page_title="Polaris Post Tracker",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# STYLES
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800;900&family=DM+Mono:wght@400;500&display=swap');

/* ── BASE ── */
html, body, .stApp {
    background: #070A12 !important;
    font-family: 'Outfit', system-ui, sans-serif !important;
    color: #CBD5E1;
}
::-webkit-scrollbar { width: 3px; background: transparent; }
::-webkit-scrollbar-thumb { background: rgba(99,102,241,.35); border-radius: 99px; }

/* ── ANIMATED BACKGROUND MESH ── */
.stApp::before {
    content: '';
    position: fixed;
    inset: 0;
    background:
        radial-gradient(ellipse 900px 600px at 15% 10%, rgba(99,102,241,.07) 0%, transparent 60%),
        radial-gradient(ellipse 700px 500px at 85% 80%, rgba(236,72,153,.05) 0%, transparent 55%),
        radial-gradient(ellipse 500px 400px at 50% 50%, rgba(6,182,212,.04) 0%, transparent 50%);
    pointer-events: none;
    z-index: 0;
}

/* ── HERO ── */
.hero {
    position: relative;
    text-align: center;
    padding: 3.5rem 2rem 2.5rem;
    margin-bottom: 0;
    overflow: hidden;
}
.hero-noise {
    position: absolute;
    inset: 0;
    background-image: url("data:image/svg+xml,%3Csvg viewBox='0 0 256 256' xmlns='http://www.w3.org/2000/svg'%3E%3Cfilter id='n'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='0.9' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23n)' opacity='0.025'/%3E%3C/svg%3E");
    opacity: .4;
    pointer-events: none;
}
.hero-rule {
    width: 48px; height: 3px;
    background: linear-gradient(90deg, #6366f1, #ec4899);
    border-radius: 99px;
    margin: 0 auto 1.4rem;
}
.hero-eyebrow {
    font-size: .72rem;
    font-weight: 600;
    letter-spacing: .22em;
    color: #6366f1;
    text-transform: uppercase;
    margin-bottom: .9rem;
    font-family: 'DM Mono', monospace;
}
.hero-title {
    font-size: 3.4rem;
    font-weight: 900;
    letter-spacing: -.05em;
    line-height: 1.05;
    margin-bottom: .5rem;
    color: #F8FAFC;
}
.hero-title span {
    background: linear-gradient(125deg, #818cf8 0%, #a78bfa 35%, #ec4899 70%, #f97316 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
}
.hero-sub {
    color: #475569;
    font-size: .9rem;
    max-width: 520px;
    margin: 0 auto 0;
    line-height: 1.6;
    font-weight: 400;
}
.hero-divider {
    height: 1px;
    background: linear-gradient(90deg, transparent, rgba(99,102,241,.2) 30%, rgba(236,72,153,.15) 70%, transparent);
    margin-top: 2.5rem;
}

/* ── SIDEBAR ── */
section[data-testid="stSidebar"] {
    background: #070A12 !important;
    border-right: 1px solid rgba(99,102,241,.1) !important;
}
section[data-testid="stSidebar"] > div { padding-top: 1.5rem !important; }
.sidebar-logo {
    display: flex; align-items: center; gap: 10px;
    padding: 0 0 1.2rem 0;
    border-bottom: 1px solid rgba(99,102,241,.1);
    margin-bottom: 1.4rem;
}
.sidebar-logo-dot {
    width: 32px; height: 32px; border-radius: 8px;
    background: linear-gradient(135deg, #6366f1, #ec4899);
    display: flex; align-items: center; justify-content: center;
    font-size: 1rem;
}
.sidebar-logo-text {
    font-size: .85rem; font-weight: 700; color: #E2E8F0; line-height: 1.2;
}
.sidebar-logo-sub { font-size: .7rem; color: #475569; font-weight: 400; }

.sidebar-section {
    margin-bottom: 1.6rem;
}
.sidebar-label {
    font-size: .65rem; font-weight: 700; letter-spacing: .15em;
    text-transform: uppercase; color: #334155;
    font-family: 'DM Mono', monospace;
    margin-bottom: .6rem; display: block;
}
.cost-chip {
    display: flex; justify-content: space-between;
    background: rgba(99,102,241,.06); border: 1px solid rgba(99,102,241,.12);
    border-radius: 8px; padding: 8px 14px; margin-bottom: 8px;
}
.cost-chip-label { font-size: .78rem; color: #64748b; }
.cost-chip-val { font-size: .78rem; font-weight: 600; color: #818cf8; font-family: 'DM Mono', monospace; }

/* ── FORM ELEMENTS ── */
div[data-baseweb="input"] > div {
    background: rgba(15,20,40,.9) !important;
    border: 1px solid rgba(99,102,241,.18) !important;
    border-radius: 10px !important;
    transition: border-color .2s ease;
}
div[data-baseweb="input"] > div:focus-within {
    border-color: rgba(99,102,241,.5) !important;
    box-shadow: 0 0 0 3px rgba(99,102,241,.1) !important;
}
div[data-baseweb="select"] > div {
    background: rgba(15,20,40,.9) !important;
    border: 1px solid rgba(99,102,241,.18) !important;
    border-radius: 10px !important;
}
input, textarea {
    color: #E2E8F0 !important;
    font-family: 'Outfit', sans-serif !important;
}

/* ── TABS ── */
.stTabs [data-baseweb="tab-list"] {
    background: rgba(15,20,40,.6) !important;
    border-radius: 12px !important;
    padding: 4px !important;
    gap: 2px !important;
    border: 1px solid rgba(99,102,241,.1) !important;
    margin-bottom: 1.5rem;
}
.stTabs [data-baseweb="tab"] {
    background: transparent !important;
    border-radius: 9px !important;
    font-size: .9rem !important;
    font-weight: 600 !important;
    padding: 9px 22px !important;
    color: #475569 !important;
    transition: all .2s ease !important;
    font-family: 'Outfit', sans-serif !important;
    border: none !important;
}
.stTabs [data-baseweb="tab"][aria-selected="true"] {
    background: rgba(99,102,241,.15) !important;
    color: #a5b4fc !important;
    box-shadow: inset 0 0 0 1px rgba(99,102,241,.25) !important;
}
.stTabs [data-baseweb="tab-highlight"] { display: none !important; }
.stTabs [data-baseweb="tab-border"] { display: none !important; }

/* ── SCRAPE BUTTON ── */
div.stButton > button {
    background: linear-gradient(135deg, #4f46e5 0%, #7c3aed 50%, #db2777 100%) !important;
    color: #fff !important;
    font-weight: 700 !important;
    font-size: .95rem !important;
    border: none !important;
    border-radius: 12px !important;
    padding: .85rem 1.5rem !important;
    box-shadow: 0 4px 24px rgba(79,70,229,.3), 0 1px 0 rgba(255,255,255,.08) inset !important;
    transition: all .3s cubic-bezier(.34,1.56,.64,1) !important;
    font-family: 'Outfit', sans-serif !important;
    letter-spacing: .01em !important;
    text-transform: none !important;
    width: 100% !important;
    position: relative !important;
    overflow: hidden !important;
}
div.stButton > button:hover {
    transform: translateY(-3px) scale(1.005) !important;
    box-shadow: 0 12px 40px rgba(79,70,229,.45), 0 1px 0 rgba(255,255,255,.1) inset !important;
}
div.stButton > button:active { transform: translateY(0) scale(.995) !important; }

/* ── DOWNLOAD BUTTON ── */
div.stDownloadButton > button {
    background: rgba(15,20,40,.7) !important;
    border: 1px solid rgba(99,102,241,.2) !important;
    color: #a5b4fc !important;
    border-radius: 10px !important;
    font-weight: 600 !important;
    font-size: .85rem !important;
    padding: .65rem 1.2rem !important;
    transition: all .2s ease !important;
    font-family: 'Outfit', sans-serif !important;
}
div.stDownloadButton > button:hover {
    border-color: rgba(99,102,241,.45) !important;
    background: rgba(99,102,241,.12) !important;
    transform: translateY(-1px) !important;
}

/* ── METRIC CARD ── */
div[data-testid="metric-container"] {
    background: rgba(15,20,40,.7) !important;
    border: 1px solid rgba(99,102,241,.12) !important;
    border-radius: 12px !important;
    padding: 1rem 1.2rem !important;
}
div[data-testid="metric-container"] label {
    font-size: .7rem !important;
    font-weight: 600 !important;
    letter-spacing: .12em !important;
    text-transform: uppercase !important;
    color: #475569 !important;
    font-family: 'DM Mono', monospace !important;
}
div[data-testid="metric-container"] [data-testid="stMetricValue"] {
    font-size: 1.8rem !important;
    font-weight: 800 !important;
    color: #F8FAFC !important;
    letter-spacing: -.03em !important;
    font-family: 'Outfit', sans-serif !important;
}

/* ── SLIDER ── */
div[data-baseweb="slider"] [role="slider"] {
    background: linear-gradient(135deg, #6366f1, #ec4899) !important;
    border: 2px solid rgba(255,255,255,.2) !important;
    box-shadow: 0 2px 8px rgba(99,102,241,.4) !important;
}

/* ── ALERTS ── */
div[data-baseweb="notification"][kind="positive"],
.stSuccess {
    background: rgba(16,185,129,.08) !important;
    border: 1px solid rgba(16,185,129,.2) !important;
    border-radius: 10px !important;
    color: #6ee7b7 !important;
}
.stError, div[data-baseweb="notification"][kind="negative"] {
    background: rgba(239,68,68,.06) !important;
    border: 1px solid rgba(239,68,68,.2) !important;
    border-radius: 10px !important;
}
.stWarning { background: rgba(245,158,11,.06) !important; border-radius: 10px !important; }

/* ── POST CARDS ── */
.glass {
    background: linear-gradient(145deg, rgba(15,22,42,.9) 0%, rgba(10,15,30,.85) 100%);
    backdrop-filter: blur(24px);
    -webkit-backdrop-filter: blur(24px);
    border: 1px solid rgba(99,102,241,.1);
    border-radius: 16px;
    padding: 22px;
    margin-bottom: 16px;
    transition: transform .3s cubic-bezier(.34,1.56,.64,1), box-shadow .3s ease, border-color .3s ease;
    box-shadow: 0 4px 24px rgba(0,0,0,.35), 0 1px 0 rgba(255,255,255,.03) inset;
    display: flex;
    flex-direction: column;
    height: 100%;
    position: relative;
    overflow: hidden;
}
.glass::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 1px;
    background: linear-gradient(90deg, transparent, rgba(99,102,241,.3), transparent);
}
.glass:hover {
    transform: translateY(-5px) scale(1.005);
    box-shadow: 0 20px 50px rgba(0,0,0,.5), 0 0 0 1px rgba(99,102,241,.2);
    border-color: rgba(99,102,241,.25);
}
.card-author {
    font-size: 1rem;
    font-weight: 700;
    color: #F1F5F9;
    margin-bottom: 2px;
    letter-spacing: -.02em;
}
.card-headline {
    font-size: .76rem;
    color: #334155;
    margin-bottom: 12px;
    line-height: 1.4;
    display: -webkit-box;
    -webkit-line-clamp: 2;
    -webkit-box-orient: vertical;
    overflow: hidden;
}
.card-snippet {
    font-size: .8rem;
    color: #7C8FA6;
    margin-bottom: 14px;
    line-height: 1.65;
    display: -webkit-box;
    -webkit-line-clamp: 4;
    -webkit-box-orient: vertical;
    overflow: hidden;
    border-left: 2px solid rgba(99,102,241,.3);
    padding-left: 12px;
    flex-grow: 1;
}
.card-date {
    display: inline-flex; align-items: center; gap: 6px;
    background: rgba(99,102,241,.1);
    color: #a5b4fc;
    padding: 4px 12px; border-radius: 999px;
    font-weight: 600; font-size: .72rem;
    border: 1px solid rgba(99,102,241,.2);
    margin-bottom: 12px; width: fit-content;
    font-family: 'DM Mono', monospace;
}
.card-date-empty {
    display: inline-flex; align-items: center; gap: 6px;
    background: rgba(51,65,85,.2); color: #475569;
    padding: 4px 12px; border-radius: 999px; font-size: .72rem;
    border: 1px solid rgba(51,65,85,.3); margin-bottom: 12px; width: fit-content;
    font-family: 'DM Mono', monospace;
}
.badge-likes {
    display: inline-flex; align-items: center; gap: 5px;
    background: rgba(236,72,153,.07); color: #f472b6;
    padding: 4px 12px; border-radius: 999px; font-weight: 600; font-size: .8rem;
    border: 1px solid rgba(236,72,153,.15); margin-bottom: 14px;
}
.card-stats {
    display: flex; gap: 8px; flex-wrap: wrap; margin-bottom: 14px;
}
.stat-pill {
    display: inline-flex; align-items: center; gap: 4px;
    background: rgba(30,41,59,.6); color: #64748b;
    padding: 3px 10px; border-radius: 6px; font-size: .72rem;
    border: 1px solid rgba(51,65,85,.4);
    font-family: 'DM Mono', monospace;
}
.card-link-linkedin {
    display: inline-flex; align-items: center; justify-content: center; gap: 7px;
    width: 100%; background: linear-gradient(135deg, #0A66C2, #0369a1);
    color: #fff !important; padding: 9px 0; border-radius: 10px;
    text-decoration: none; font-weight: 600; font-size: .82rem;
    transition: all .25s ease; margin-bottom: 8px;
    box-shadow: 0 2px 12px rgba(10,102,194,.25);
}
.card-link-linkedin:hover { filter: brightness(1.12); transform: scale(1.02); box-shadow: 0 6px 20px rgba(10,102,194,.4); }
.card-link-x {
    display: inline-flex; align-items: center; justify-content: center; gap: 7px;
    width: 100%; background: rgba(15,23,42,.8);
    color: #E2E8F0 !important; padding: 9px 0; border-radius: 10px;
    border: 1px solid rgba(99,102,241,.15);
    text-decoration: none; font-weight: 600; font-size: .82rem;
    transition: all .25s ease; margin-bottom: 8px;
}
.card-link-x:hover { border-color: rgba(99,102,241,.4); background: rgba(99,102,241,.1); transform: scale(1.02); }
.card-ts {
    font-size: .64rem; color: #1E293B; margin-top: auto;
    padding-top: 10px; border-top: 1px solid rgba(255,255,255,.04);
    font-family: 'DM Mono', monospace;
}

/* ── SECTION HEADER ── */
.section-header {
    display: flex; align-items: center; justify-content: space-between;
    margin-bottom: 1.2rem; margin-top: .5rem;
}
.section-title {
    font-size: 1rem; font-weight: 700; color: #E2E8F0;
    display: flex; align-items: center; gap: 8px; letter-spacing: -.02em;
}
.section-count {
    background: rgba(99,102,241,.12); color: #818cf8;
    padding: 3px 10px; border-radius: 6px; font-size: .75rem;
    font-weight: 600; font-family: 'DM Mono', monospace;
    border: 1px solid rgba(99,102,241,.2);
}

/* ── SUCCESS BANNER ── */
.success-banner {
    background: linear-gradient(135deg, rgba(99,102,241,.08), rgba(236,72,153,.06));
    border: 1px solid rgba(99,102,241,.2);
    border-radius: 12px; padding: 14px 20px;
    font-size: .88rem; color: #a5b4fc; margin-bottom: 18px;
    display: flex; align-items: center; gap: 10px;
}

/* ── METRIC ROW ── */
.metric-row { display: flex; gap: 10px; margin-bottom: 20px; flex-wrap: wrap; }
.metric-card {
    background: linear-gradient(145deg, rgba(15,22,42,.8), rgba(10,15,30,.7));
    border: 1px solid rgba(99,102,241,.1);
    border-radius: 12px; padding: 14px 18px; flex: 1; min-width: 110px;
    position: relative; overflow: hidden;
}
.metric-card::after {
    content: ''; position: absolute; top: 0; left: 0; right: 0; height: 2px;
    background: linear-gradient(90deg, #6366f1, #ec4899);
    opacity: .6;
}
.metric-label {
    font-size: .65rem; color: #475569; text-transform: uppercase;
    letter-spacing: .1em; margin-bottom: 6px;
    font-family: 'DM Mono', monospace; font-weight: 500;
}
.metric-value { font-size: 1.7rem; font-weight: 800; color: #F1F5F9; letter-spacing: -.04em; }

/* ── DIVIDER ── */
hr, .stDivider {
    border-color: rgba(99,102,241,.1) !important;
    margin: 1.2rem 0 !important;
}

/* ── TOGGLE ── */
div[data-baseweb="checkbox"] span {
    font-size: .85rem !important;
    color: #94A3B8 !important;
    font-family: 'Outfit', sans-serif !important;
}

/* ── EXPANDER ── */
details summary {
    background: rgba(15,20,40,.7) !important;
    border-radius: 10px !important;
    color: #94A3B8 !important;
}

/* ── PROGRESS ── */
div[data-testid="stProgress"] > div > div {
    background: linear-gradient(90deg, #6366f1, #ec4899) !important;
    border-radius: 99px !important;
}
div[data-testid="stProgress"] > div {
    background: rgba(30,41,59,.5) !important;
    border-radius: 99px !important;
}

/* ── LABEL ── */
label[data-testid="stWidgetLabel"] p, .stLabel {
    font-size: .8rem !important;
    font-weight: 600 !important;
    color: #64748B !important;
    letter-spacing: .03em !important;
    text-transform: none !important;
    font-family: 'Outfit', sans-serif !important;
}
</style>
""", unsafe_allow_html=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# TIMESTAMP PARSER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _parse_timestamp(item):
    for key in ("posted_at", "createdAt", "created_at", "postedAtISO", "timeSincePosted", "publishedAt"):
        val = item.get(key)
        if val is None:
            continue
        if isinstance(val, dict):
            ts_ms = val.get("timestamp")
            if ts_ms is not None:
                try:
                    ts = int(ts_ms)
                    if ts > 1_000_000_000_000:
                        ts //= 1000
                    return datetime.fromtimestamp(ts, tz=timezone.utc)
                except (ValueError, OSError):
                    pass
            date_str = val.get("date", "")
            if date_str:
                for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
                    try:
                        return datetime.strptime(date_str, fmt).replace(tzinfo=timezone.utc)
                    except ValueError:
                        pass
            continue
        raw = str(val).strip()
        if not raw:
            continue
        if raw.isdigit():
            ts = int(raw)
            if ts > 1_000_000_000_000:
                ts //= 1000
            try:
                return datetime.fromtimestamp(ts, tz=timezone.utc)
            except (ValueError, OSError):
                pass
            continue
        for fmt in (
            "%Y-%m-%dT%H:%M:%S.%fZ",
            "%Y-%m-%dT%H:%M:%SZ",
            "%a %b %d %H:%M:%S %z %Y",
            "%Y-%m-%dT%H:%M:%S",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d",
        ):
            try:
                dt = datetime.strptime(raw, fmt)
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                return dt
            except ValueError:
                pass
    return None


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# TIME FILTER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _filter_by_time(posts, time_period, custom_dates):
    if time_period == "today":
        midnight = datetime.now(IST).replace(hour=0, minute=0, second=0, microsecond=0)
        return [p for p in posts if p["PostedDT"] is None or p["PostedDT"] >= midnight]
    if time_period == "custom" and custom_dates and len(custom_dates) > 0:
        sd = custom_dates[0]
        ed = custom_dates[1] if len(custom_dates) > 1 else custom_dates[0]
        # Handle both date and datetime objects
        if isinstance(sd, datetime):
            start = sd.replace(tzinfo=IST) if sd.tzinfo is None else sd
        else:
            start = datetime(sd.year, sd.month, sd.day, 0, 0, 0, tzinfo=IST)
        if isinstance(ed, datetime):
            end = ed.replace(tzinfo=IST) if ed.tzinfo is None else ed
        else:
            end = datetime(ed.year, ed.month, ed.day, 23, 59, 59, tzinfo=IST)
        return [p for p in posts if p["PostedDT"] and start <= p["PostedDT"] <= end]
    return posts


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# DATA INGESTION — LINKEDIN
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _ingest_linkedin(raw_items, keyword, time_period, custom_dates):
    posts = []
    kw_lower = keyword.lower()

    for item in raw_items:
        if not isinstance(item, dict):
            continue

        # ── Author ──
        af = item.get("author")
        if isinstance(af, dict):
            first = af.get("firstName", "")
            last = af.get("lastName", "")
            full = f"{first} {last}".strip()
            author = af.get("name", "").strip() or full or str(item.get("authorName", "")).strip() or "Unknown"
            headline = af.get("headline", "").strip() or str(item.get("authorHeadline", "")).strip()
            author_img = af.get("picture", "") or af.get("image_url", "") or str(item.get("authorProfilePicture", "")).strip()
            profile_url = af.get("url", "") or af.get("profileUrl", "") or str(item.get("authorProfileUrl", "")).strip()
        else:
            author = (str(af).strip() if af else str(item.get("authorName", "")).strip()) or "Unknown"
            headline = str(item.get("authorHeadline", "")).strip()
            author_img = str(item.get("authorProfilePicture", "")).strip()
            profile_url = str(item.get("authorProfileUrl", "")).strip()

        # ── Reactions ──
        stats = item.get("stats")
        if isinstance(stats, dict):
            likes = int(stats.get("total_reactions", 0) or 0)
            comments = int(stats.get("numComments", 0) or stats.get("comments", 0) or 0)
            reposts = int(stats.get("numShares", 0) or stats.get("shares", 0) or 0)
        else:
            likes = 0
            comments = 0
            reposts = 0
            for k in ("likes", "numLikes", "reactionCount"):
                v = item.get(k)
                if v is not None:
                    try:
                        likes = int(v)
                        break
                    except (ValueError, TypeError):
                        pass
            for k in ("numComments", "comments", "commentCount"):
                v = item.get(k)
                if v is not None:
                    try:
                        comments = int(v)
                        break
                    except (ValueError, TypeError):
                        pass
            for k in ("numShares", "shares", "reshareCount", "reposts"):
                v = item.get(k)
                if v is not None:
                    try:
                        reposts = int(v)
                        break
                    except (ValueError, TypeError):
                        pass

        # ── Post URL ──
        activity_id = str(item.get("activity_id", "")).strip()
        post_url = str(item.get("post_url", "") or "").strip()
        if not post_url:
            for k in ("url", "postUrl", "link", "permalink"):
                if item.get(k):
                    post_url = str(item[k]).strip()
                    break
        if not post_url and activity_id and activity_id.isdigit():
            post_url = f"https://www.linkedin.com/feed/update/urn:li:activity:{activity_id}/"
        if not post_url:
            continue

        # ── Timestamp ──
        posted_dt = _parse_timestamp(item)
        if posted_dt:
            posted_date = posted_dt.astimezone(IST).strftime("%d %b %Y")
            posted_time = posted_dt.astimezone(IST).strftime("%I:%M %p IST")
            posted_datetime_str = posted_dt.astimezone(IST).strftime("%d %b %Y · %I:%M %p IST")
        else:
            pa = item.get("posted_at") or item.get("postedAtISO") or item.get("timeSincePosted")
            if isinstance(pa, dict) and pa.get("display_text"):
                posted_datetime_str = pa["display_text"]
            else:
                posted_datetime_str = str(pa) if pa else ""
            posted_date = ""
            posted_time = ""

        # ── Text ──
        raw_text = str(item.get("text", "")).strip()
        snippet = raw_text[:300] + ("…" if len(raw_text) > 300 else "")

        # ── Impressions estimate (reactions × 80, floor at 0) ──
        impressions_est = likes * 80

        # ── Tags Polaris check ──
        tags_polaris = "Yes" if (
            "polaris" in raw_text.lower() or
            "polariscampus" in raw_text.lower() or
            "polaris school" in raw_text.lower() or
            kw_lower in raw_text.lower()
        ) else "No"

        posts.append({
            "Date": posted_date,
            "Time": posted_time,
            "DateTime (IST)": posted_datetime_str,
            "Account Name": author,
            "Headline / Bio": headline,
            "Profile URL": profile_url,
            "Post Link": post_url,
            "Post Text (Preview)": snippet,
            "Reactions": likes,
            "Comments": comments,
            "Reposts": reposts,
            "Est. Impressions": impressions_est,
            "Tags Polaris": tags_polaris,
            "Platform": "LinkedIn",
            "Scraped At": datetime.now(IST).strftime("%d %b %Y %H:%M IST"),
            # Internal fields
            "ActivityID": activity_id,
            "AuthorImg": author_img,
            "PostedDT": posted_dt,
        })

    # Deduplicate
    seen = set()
    unique = []
    for p in posts:
        aid = p["ActivityID"]
        if aid and aid in seen:
            continue
        if aid:
            seen.add(aid)
        unique.append(p)

    unique = _filter_by_time(unique, time_period, custom_dates)
    unique.sort(key=lambda p: p["PostedDT"] or datetime.min.replace(tzinfo=timezone.utc), reverse=True)
    return unique


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# DATA INGESTION — X / TWITTER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _ingest_x(raw_items, keyword, time_period, custom_dates):
    posts = []
    kw_lower = keyword.lower()

    for item in raw_items:
        if not isinstance(item, dict):
            continue

        ai = item.get("author") or item.get("user") or {}
        if not isinstance(ai, dict):
            ai = {}

        author = ai.get("name") or item.get("authorName") or item.get("userName") or "Unknown"
        handle = (
            ai.get("username") or ai.get("userName") or ai.get("screen_name")
            or item.get("userHandle") or ""
        )
        if handle and not handle.startswith("@"):
            handle = f"@{handle}"

        bio = ai.get("description") or ai.get("bio") or item.get("authorBio") or ""
        author_img = (
            ai.get("profilePicture") or ai.get("profile_image_url_https")
            or ai.get("avatar") or item.get("authorAvatar") or ""
        )
        profile_url = f"https://x.com/{handle.lstrip('@')}" if handle else ""

        likes = item.get("likes") or item.get("likeCount") or item.get("favorite_count") or 0
        try:
            likes = int(likes)
        except (ValueError, TypeError):
            likes = 0

        replies = int(item.get("replies") or item.get("reply_count") or 0)
        retweets = int(item.get("retweets") or item.get("retweet_count") or 0)

        url = item.get("url") or item.get("tweetUrl") or ""
        if not url and handle and item.get("id"):
            url = f"https://x.com/{handle.lstrip('@')}/status/{item['id']}"

        text = item.get("text") or item.get("full_text") or item.get("tweetText") or ""
        snippet = text[:300] + ("…" if len(text) > 300 else "")

        posted_dt = _parse_timestamp(item)
        if posted_dt:
            posted_date = posted_dt.astimezone(IST).strftime("%d %b %Y")
            posted_time = posted_dt.astimezone(IST).strftime("%I:%M %p IST")
            posted_datetime_str = posted_dt.astimezone(IST).strftime("%d %b %Y · %I:%M %p IST")
        else:
            posted_datetime_str = str(item.get("createdAt") or item.get("created_at") or "")
            posted_date = ""
            posted_time = ""

        tweet_id = str(item.get("id", item.get("tweetId", "")))
        impressions_est = likes * 35

        tags_polaris = "Yes" if (
            "polaris" in text.lower() or
            "polariscampus" in text.lower() or
            kw_lower in text.lower()
        ) else "No"

        posts.append({
            "Date": posted_date,
            "Time": posted_time,
            "DateTime (IST)": posted_datetime_str,
            "Account Name": f"{author} ({handle})" if handle else author,
            "Headline / Bio": bio,
            "Profile URL": profile_url,
            "Post Link": url,
            "Post Text (Preview)": snippet,
            "Reactions": likes,
            "Comments": replies,
            "Reposts": retweets,
            "Est. Impressions": impressions_est,
            "Tags Polaris": tags_polaris,
            "Platform": "X (Twitter)",
            "Scraped At": datetime.now(IST).strftime("%d %b %Y %H:%M IST"),
            "ActivityID": tweet_id,
            "AuthorImg": author_img,
            "PostedDT": posted_dt,
        })

    seen = set()
    unique = []
    for p in posts:
        aid = p["ActivityID"]
        if aid and aid in seen:
            continue
        if aid:
            seen.add(aid)
        unique.append(p)

    unique = _filter_by_time(unique, time_period, custom_dates)
    unique.sort(key=lambda p: p["PostedDT"] or datetime.min.replace(tzinfo=timezone.utc), reverse=True)
    return unique


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# EXCEL EXPORT — BRANDED & FORMATTED
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _build_excel(posts_linkedin, posts_x, keyword):
    buf = io.BytesIO()

    EXPORT_COLS = [
        "Date", "Time", "DateTime (IST)", "Account Name", "Headline / Bio",
        "Profile URL", "Post Link", "Post Text (Preview)",
        "Reactions", "Comments", "Reposts", "Est. Impressions",
        "Tags Polaris", "Platform", "Scraped At"
    ]

    # Combine all posts
    all_posts = posts_linkedin + posts_x
    all_posts.sort(key=lambda p: p.get("PostedDT") or datetime.min.replace(tzinfo=timezone.utc), reverse=True)
    polaris_posts = [p for p in all_posts if p.get("Tags Polaris") == "Yes"]

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # ── Sheet 1: All Posts ──
        if all_posts:
            df_all = pd.DataFrame(all_posts)[EXPORT_COLS]
        else:
            df_all = pd.DataFrame(columns=EXPORT_COLS)
        df_all.to_excel(writer, sheet_name="All Posts", index=False)

        # ── Sheet 2: Tags Polaris Only ──
        if polaris_posts:
            df_pol = pd.DataFrame(polaris_posts)[EXPORT_COLS]
        else:
            df_pol = pd.DataFrame(columns=EXPORT_COLS)
        df_pol.to_excel(writer, sheet_name="Tags Polaris", index=False)

        # ── Sheet 3: LinkedIn Only ──
        if posts_linkedin:
            df_li = pd.DataFrame(posts_linkedin)[EXPORT_COLS]
        else:
            df_li = pd.DataFrame(columns=EXPORT_COLS)
        df_li.to_excel(writer, sheet_name="LinkedIn", index=False)

        # ── Sheet 4: X Only ──
        if posts_x:
            df_x = pd.DataFrame(posts_x)[EXPORT_COLS]
        else:
            df_x = pd.DataFrame(columns=EXPORT_COLS)
        df_x.to_excel(writer, sheet_name="X (Twitter)", index=False)

        wb = writer.book

        # Style each sheet
        HEADER_FILL = PatternFill("solid", fgColor="1a0533")   # deep purple
        HEADER_FONT = Font(bold=True, color="C4B5FD", size=10, name="Calibri")
        POLARIS_FILL = PatternFill("solid", fgColor="1e0a3c")   # slight purple tint for polaris rows
        ALT_FILL = PatternFill("solid", fgColor="0d1424")       # dark blue-grey
        BASE_FILL = PatternFill("solid", fgColor="080f1a")      # near black
        YES_FONT = Font(bold=True, color="A78BFA", size=9, name="Calibri")
        NO_FONT = Font(color="475569", size=9, name="Calibri")
        THIN_BORDER = Border(
            bottom=Side(style="thin", color="1e293b"),
        )

        COL_WIDTHS = {
            "Date": 14, "Time": 13, "DateTime (IST)": 24,
            "Account Name": 28, "Headline / Bio": 35,
            "Profile URL": 40, "Post Link": 40,
            "Post Text (Preview)": 55,
            "Reactions": 12, "Comments": 12, "Reposts": 12,
            "Est. Impressions": 18, "Tags Polaris": 14,
            "Platform": 14, "Scraped At": 22
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.sheet_view.showGridLines = False
            ws.sheet_properties.tabColor = "7C3AED"

            # Header row
            for col_idx, col_name in enumerate(EXPORT_COLS, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
                # Set column width
                width = COL_WIDTHS.get(col_name, 18)
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            ws.row_dimensions[1].height = 22

            # Data rows
            for row_idx in range(2, ws.max_row + 1):
                is_alt = (row_idx % 2 == 0)
                for col_idx in range(1, len(EXPORT_COLS) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    col_name = EXPORT_COLS[col_idx - 1]
                    # Tags Polaris column
                    if col_name == "Tags Polaris":
                        if str(cell.value) == "Yes":
                            cell.font = YES_FONT
                            cell.fill = POLARIS_FILL
                        else:
                            cell.font = NO_FONT
                            cell.fill = ALT_FILL if is_alt else BASE_FILL
                    else:
                        cell.fill = ALT_FILL if is_alt else BASE_FILL
                        cell.font = Font(color="94A3B8", size=9, name="Calibri")
                        if col_name in ("Reactions", "Comments", "Reposts", "Est. Impressions"):
                            cell.alignment = Alignment(horizontal="right")
                        else:
                            cell.alignment = Alignment(horizontal="left", wrap_text=False, vertical="center")
                    cell.border = THIN_BORDER
                ws.row_dimensions[row_idx].height = 18

            # Freeze top row
            ws.freeze_panes = "A2"

    buf.seek(0)
    return buf.getvalue()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SCRAPING ENGINE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _run_scrape(platform, keyword, time_period, custom_dates, token, max_posts, debug):
    label = "LinkedIn" if platform == "linkedin" else "X (Twitter)"

    with st.status(f"🔄 Scraping {label}…", expanded=True) as status_ui:
        try:
            if platform == "linkedin":
                run_url = f"{APIFY_BASE}/acts/{ACTOR_LINKEDIN}/runs?token={token}"
                kw_enc = urllib.parse.quote(keyword.strip())
                search_url = f"https://www.linkedin.com/search/results/content/?keywords={kw_enc}&sortBy=date_posted"
                if time_period in ("today", "past-24h", "past-week", "past-month"):
                    api_filter = "past-24h" if time_period == "today" else time_period
                    search_url += f'&datePosted="{api_filter}"'
                payload = {
                    "urls": [search_url],
                    "startUrls": [{"url": search_url}],
                    "searchKeywords": keyword.strip(),
                    "maxItems": max_posts,
                }
            else:
                run_url = f"{APIFY_BASE}/acts/{ACTOR_X}/runs?token={token}"
                final_kw = keyword.strip()
                if time_period == "custom" and custom_dates and len(custom_dates) > 0:
                    sd = custom_dates[0]
                    ed = custom_dates[1] if len(custom_dates) > 1 else custom_dates[0]
                    sd_str = sd.strftime("%Y-%m-%d") if isinstance(sd, datetime) else sd.isoformat()
                    ed_obj = (ed + timedelta(days=1))
                    ed_str = ed_obj.strftime("%Y-%m-%d") if isinstance(ed_obj, datetime) else ed_obj.isoformat()
                    final_kw += f" since:{sd_str} until:{ed_str}"
                payload = {
                    "searchTerms": [final_kw],
                    "tweetsToScrape": max_posts,
                    "maxItems": max_posts,
                }
                if time_period in ("past-24h", "past-week", "past-month"):
                    payload["timePeriod"] = time_period
                elif time_period == "today":
                    payload["timePeriod"] = "past-24h"

            st.write("🚀 Starting Apify actor run…")
            if debug:
                st.json(payload)

            resp = requests.post(run_url, json=payload, timeout=30)
            if resp.status_code not in (200, 201):
                st.error(f"❌ API Error {resp.status_code}: {resp.text[:500]}")
                status_ui.update(label="❌ Failed to start", state="error")
                return

            run_data = resp.json().get("data", {})
            run_id = run_data.get("id", "")
            dataset_id = run_data.get("defaultDatasetId", "")

            if not run_id:
                st.error("❌ No run ID returned from Apify. Check your API token.")
                status_ui.update(label="❌ Failed", state="error")
                return

            st.write(f"✅ Run started · ID: `{run_id}`")

            progress = st.progress(0)
            status_line = st.empty()

            for i in range(150):
                time.sleep(2)
                elapsed = (i + 1) * 2
                progress.progress(min(int((elapsed / 300) * 100), 95))
                try:
                    check = requests.get(f"{APIFY_BASE}/actor-runs/{run_id}?token={token}", timeout=30)
                    info = check.json().get("data", {})
                    run_status = info.get("status", "UNKNOWN")
                    ds_id = info.get("defaultDatasetId") or dataset_id
                except Exception as e:
                    status_line.warning(f"⚠️ Poll error: {e}")
                    continue

                status_line.info(f"⏳ Status: **{run_status}** — {elapsed}s elapsed")

                if run_status == "SUCCEEDED":
                    progress.progress(100)
                    st.write("📦 Fetching results…")
                    items_resp = requests.get(
                        f"{APIFY_BASE}/datasets/{ds_id}/items?token={token}&format=json",
                        timeout=120,
                    )
                    if items_resp.status_code != 200:
                        st.error(f"❌ Dataset fetch failed: {items_resp.status_code}")
                        status_ui.update(label="❌ Dataset error", state="error")
                        return

                    raw_items = items_resp.json()
                    st.write(f"📊 Raw items received: **{len(raw_items)}**")
                    if debug and raw_items:
                        st.json(raw_items[0])

                    new_posts = (
                        _ingest_linkedin(raw_items, keyword, time_period, custom_dates)
                        if platform == "linkedin"
                        else _ingest_x(raw_items, keyword, time_period, custom_dates)
                    )

                    # Accumulate mode
                    if st.session_state.get("accumulate_mode", False):
                        existing = st.session_state.get(f"posts_{platform}", [])
                        if existing:
                            new_aids = {p["ActivityID"] for p in new_posts if p["ActivityID"]}
                            for old_p in existing:
                                old_aid = old_p.get("ActivityID", "")
                                if old_aid and old_aid not in new_aids:
                                    new_posts.append(old_p)
                                    new_aids.add(old_aid)
                            new_posts.sort(
                                key=lambda p: p["PostedDT"] or datetime.min.replace(tzinfo=timezone.utc),
                                reverse=True,
                            )
                            st.write(f"📚 Accumulated: **{len(new_posts)}** unique posts")

                    st.session_state[f"posts_{platform}"] = new_posts
                    st.session_state[f"last_keyword_{platform}"] = keyword.strip()
                    st.session_state[f"last_period_{platform}"] = time_period
                    st.session_state[f"last_dates_{platform}"] = custom_dates
                    st.session_state[f"scraped_at_{platform}"] = datetime.now(IST).strftime("%d %b %Y %H:%M IST")
                    status_ui.update(label=f"✅ Found {len(new_posts)} posts!", state="complete")
                    st.balloons()
                    return

                if run_status in ("FAILED", "ABORTED", "TIMED-OUT"):
                    st.error(f"❌ Run ended with status: **{run_status}**.")
                    if debug:
                        st.json(info)
                    status_ui.update(label=f"❌ {run_status}", state="error")
                    return

            st.error("⏱ Timed out after 5 minutes. Try fewer posts or a narrower date window.")
            status_ui.update(label="⏱ Timed out", state="error")

        except Exception as e:
            st.error(f"❌ Unexpected error: {e}")
            status_ui.update(label="❌ Error", state="error")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# RENDER RESULTS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _render_results(platform):
    posts = st.session_state.get(f"posts_{platform}", [])
    if not posts:
        return

    kw = st.session_state.get(f"last_keyword_{platform}", "")
    per = st.session_state.get(f"last_period_{platform}", "")
    dates = st.session_state.get(f"last_dates_{platform}")
    scraped_at = st.session_state.get(f"scraped_at_{platform}", "")

    now_utc = datetime.now(timezone.utc)
    midnight_ist = datetime.now(IST).replace(hour=0, minute=0, second=0, microsecond=0)

    cnt_today = sum(1 for p in posts if p.get("PostedDT") and p["PostedDT"] >= midnight_ist)
    total_reactions = sum(p.get("Reactions", 0) for p in posts)
    total_impressions = sum(p.get("Est. Impressions", 0) for p in posts)
    polaris_count = sum(1 for p in posts if p.get("Tags Polaris") == "Yes")

    label = "LinkedIn" if platform == "linkedin" else "X (Twitter)"
    per_str = per
    if per == "custom" and dates and len(dates) > 0:
        if len(dates) == 1 or (len(dates) > 1 and dates[0] == dates[1]):
            d = dates[0]
            per_str = d.strftime("%d %b %Y") if isinstance(d, datetime) else str(d)
        else:
            d0, d1 = dates[0], dates[1]
            s0 = d0.strftime("%d %b") if isinstance(d0, datetime) else str(d0)
            s1 = d1.strftime("%d %b %Y") if isinstance(d1, datetime) else str(d1)
            per_str = f"{s0} – {s1}"

    st.markdown(
        f'<div class="success-banner">✅ Found <strong>{len(posts)}</strong> posts on {label} '
        f'for <strong>"{kw}"</strong> · {per_str} &mdash; scraped {scraped_at}</div>',
        unsafe_allow_html=True
    )

    # Metrics
    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("📋 Total Posts", len(posts))
    m2.metric("📅 Today", cnt_today)
    m3.metric("🎯 Tags Polaris", polaris_count)
    m4.metric("❤️ Total Reactions", f"{total_reactions:,}")
    m5.metric("👀 Est. Impressions", f"{total_impressions:,}")

    st.caption("Est. Impressions = Reactions × 80 (LinkedIn) or × 35 (X). Sample only, not full universe.")
    st.markdown("")

    # Table toggle
    if st.toggle(f"📄 Show raw data table", value=False, key=f"raw_{platform}"):
        DISPLAY_COLS = ["Date", "Account Name", "Post Link", "Reactions", "Comments", "Reposts", "Est. Impressions", "Tags Polaris"]
        df_show = pd.DataFrame(posts)[[c for c in DISPLAY_COLS if c in pd.DataFrame(posts).columns]]
        st.dataframe(
            df_show, use_container_width=True, hide_index=True,
            column_config={"Post Link": st.column_config.LinkColumn("Post Link")},
        )

    # Filter toggle
    show_polaris_only = st.toggle(f"🎯 Show only posts tagging Polaris", value=False, key=f"pol_{platform}")
    display_posts = [p for p in posts if p.get("Tags Polaris") == "Yes"] if show_polaris_only else posts

    st.markdown(f'<div class="section-title">🏆 {label} Posts ({len(display_posts)} shown)</div>', unsafe_allow_html=True)

    def _esc(s):
        return html.escape(str(s or ""), quote=True)

    for row_start in range(0, len(display_posts), 3):
        cols = st.columns(3)
        for j in range(3):
            idx = row_start + j
            if idx >= len(display_posts):
                break
            p = display_posts[idx]

            author = _esc(p.get("Account Name", ""))
            headline_text = _esc(p.get("Headline / Bio", ""))
            snippet = _esc(p.get("Post Text (Preview)", ""))
            posted = _esc(p.get("DateTime (IST)", ""))
            post_link = _esc(p.get("Post Link", ""))
            scraped_ts = _esc(p.get("Scraped At", ""))
            tags_pol = p.get("Tags Polaris", "No")

            if p.get("AuthorImg"):
                img_html = (
                    f'<img src="{_esc(p["AuthorImg"])}" alt="" '
                    f'style="width:42px;height:42px;border-radius:50%;object-fit:cover;'
                    f'margin-right:10px;flex-shrink:0;border:2px solid rgba(139,92,246,.3);" />'
                )
            else:
                initial = author[0].upper() if author else "?"
                img_html = (
                    f'<div style="width:42px;height:42px;border-radius:50%;'
                    f'background:linear-gradient(135deg,#7c3aed,#4f46e5);display:flex;'
                    f'align-items:center;justify-content:center;font-size:1.1rem;'
                    f'font-weight:700;color:#fff;margin-right:10px;flex-shrink:0;">'
                    f'{initial}</div>'
                )

            pol_badge = (
                '<span style="background:rgba(139,92,246,.15);color:#a78bfa;'
                'border:1px solid rgba(139,92,246,.3);border-radius:6px;'
                'padding:2px 8px;font-size:.7rem;font-weight:600;margin-left:6px;">'
                '🎯 Tags Polaris</span>'
            ) if tags_pol == "Yes" else ""

            hl = f'<div class="card-headline">{headline_text}</div>' if headline_text else ""
            snip = f'<div class="card-snippet">{snippet}</div>' if snippet else ""
            date_badge = (
                f'<div class="card-date">📅 {posted}</div>' if posted
                else '<div class="card-date-empty">📅 No date</div>'
            )

            btn_class = "card-link-x" if platform == "x" else "card-link-linkedin"
            pf_name = "X" if platform == "x" else "LinkedIn"

            with cols[j]:
                st.markdown(
                    f'<div class="glass">'
                    f'<div style="display:flex;align-items:center;margin-bottom:8px;">'
                    f'{img_html}'
                    f'<div><div class="card-author">{author}{pol_badge}</div>{hl}</div>'
                    f'</div>'
                    f'{snip}{date_badge}'
                    f'<div class="badge-likes">❤️ {int(p.get("Reactions", 0)):,} Reactions</div>'
                    f'<a href="{post_link}" target="_blank" rel="noopener noreferrer" '
                    f'class="{btn_class}">🔗&nbsp; View on {pf_name}</a>'
                    f'<div class="card-ts">🕒 Scraped {scraped_ts}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

    # Downloads
    st.markdown("---")
    dl1, dl2, _ = st.columns([1, 1, 3])
    EXPORT_COLS_DISPLAY = [
        "Date", "Time", "DateTime (IST)", "Account Name", "Headline / Bio",
        "Profile URL", "Post Link", "Post Text (Preview)",
        "Reactions", "Comments", "Reposts", "Est. Impressions",
        "Tags Polaris", "Platform", "Scraped At"
    ]
    df_export = pd.DataFrame(posts)
    df_export_cols = [c for c in EXPORT_COLS_DISPLAY if c in df_export.columns]
    df_export = df_export[df_export_cols]

    dl1.download_button(
        "📥 Download CSV",
        data=df_export.to_csv(index=False).encode("utf-8"),
        file_name=f"polaris_{platform}_{kw.replace(' ', '_')}.csv",
        mime="text/csv",
        use_container_width=True,
        key=f"dl_csv_{platform}",
    )

    # Excel with formatting
    try:
        posts_li = st.session_state.get("posts_linkedin", []) if platform == "linkedin" else []
        posts_xp = st.session_state.get("posts_x", []) if platform == "x" else []
        if platform == "linkedin":
            xl_data = _build_excel(posts, [], kw)
        else:
            xl_data = _build_excel([], posts, kw)
        dl2.download_button(
            "📥 Download Excel (4 Sheets)",
            data=xl_data,
            file_name=f"polaris_{platform}_{kw.replace(' ', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=f"dl_xls_{platform}",
        )
    except Exception as exc:
        dl2.error(f"Excel export error: {exc}")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# COMBINED EXCEL DOWNLOAD (BOTH PLATFORMS)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _render_combined_download():
    posts_li = st.session_state.get("posts_linkedin", [])
    posts_xp = st.session_state.get("posts_x", [])
    if not posts_li and not posts_xp:
        return

    st.markdown("---")
    st.markdown("### 📦 Combined Export (Both Platforms)")
    kw = st.session_state.get("last_keyword_linkedin", "") or st.session_state.get("last_keyword_x", "") or "polaris"
    total = len(posts_li) + len(posts_xp)
    pol_count = sum(1 for p in posts_li + posts_xp if p.get("Tags Polaris") == "Yes")
    st.caption(f"{total} total posts across both platforms · {pol_count} tag Polaris")

    try:
        xl_data = _build_excel(posts_li, posts_xp, kw)
        st.download_button(
            f"📥 Download Combined Excel — {total} posts · {pol_count} tag Polaris",
            data=xl_data,
            file_name=f"polaris_all_platforms_{datetime.now(IST).strftime('%d%b%Y')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=False,
        )
        st.success("✅ Excel ready to download — 4 tabs: All Posts · Tags Polaris · LinkedIn · X (Twitter)")
    except Exception as exc:
        st.error(f"Combined export error: {exc}")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SESSION INIT
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
for _key in ("posts_linkedin", "posts_x"):
    if _key not in st.session_state:
        st.session_state[_key] = []
if "accumulate_mode" not in st.session_state:
    st.session_state.accumulate_mode = False


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# HERO
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
st.markdown("""
<div class="hero">
    <div class="hero-noise"></div>
    <div class="hero-eyebrow">Polaris School of Technology &nbsp;·&nbsp; Brand Intelligence</div>
    <div class="hero-rule"></div>
    <div class="hero-title">Post <span>Tracker</span></div>
    <div class="hero-sub">Track every LinkedIn &amp; X mention of Polaris in real time · Filter by date · Export branded Excel reports</div>
    <div class="hero-divider"></div>
</div>
""", unsafe_allow_html=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SIDEBAR
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
with st.sidebar:
    st.markdown("""
    <div class="sidebar-logo">
        <div class="sidebar-logo-dot">🎯</div>
        <div>
            <div class="sidebar-logo-text">Post Tracker</div>
            <div class="sidebar-logo-sub">Polaris Brand Intelligence</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<span class="sidebar-label">API Configuration</span>', unsafe_allow_html=True)
    api_token = st.text_input(
        "Apify API Token",
        value=os.getenv("APIFY_TOKEN", ""),
        type="password",
        placeholder="apify_api_xxxxxxxxxxxx",
        help="Get your token from apify.com → Settings → Integrations",
        label_visibility="collapsed",
    )
    if api_token:
        st.success("✅ Token connected")
    else:
        st.warning("Enter your Apify API token")

    st.markdown("---")
    st.markdown('<span class="sidebar-label">Run Settings</span>', unsafe_allow_html=True)
    debug_mode = st.toggle("Debug Mode", value=False)
    accumulate_mode = st.toggle(
        "Accumulate Runs",
        value=False,
        help="ON: merges new results with previous scrapes. OFF: fresh slate each run.",
    )

    st.markdown("---")
    st.markdown('<span class="sidebar-label">Estimated Cost</span>', unsafe_allow_html=True)
    st.markdown("""
    <div class="cost-chip">
        <span class="cost-chip-label">LinkedIn</span>
        <span class="cost-chip-val">$0.005 / post</span>
    </div>
    <div class="cost-chip">
        <span class="cost-chip-label">X / Twitter</span>
        <span class="cost-chip-val">$0.00015 / tweet</span>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("---")
    st.markdown('<span class="sidebar-label">Deployment</span>', unsafe_allow_html=True)
    st.caption("Hosted on Streamlit Community Cloud")
    st.markdown("[Deploy guide ↗](https://share.streamlit.io)")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# TABS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
tab_li, tab_x, tab_help = st.tabs(["  LinkedIn  ", "  X (Twitter)  ", "  How to Use  "])

# ━━ TAB 1: LINKEDIN ━━
with tab_li:
    col_kw, col_tf = st.columns([2.5, 1])
    with col_kw:
        keyword_li = st.text_input(
            "Keyword / Brand / Hashtag",
            value="Polaris School of Technology",
            key="kw_li",
        )
    with col_tf:
        period_li = st.selectbox(
            "Time Window",
            ["today", "past-24h", "past-week", "past-month", "custom"],
            key="pd_li",
        )

    dates_li = None
    if period_li == "custom":
        dates_li = st.date_input(
            "📅 Date Range (Start → End)",
            value=(datetime.today().date() - timedelta(days=7), datetime.today().date()),
            max_value=datetime.today().date(),
            key="dates_li",
        )
        # Normalize to list
        if isinstance(dates_li, (list, tuple)):
            dates_li = list(dates_li)
        else:
            dates_li = [dates_li]

    col_slider, col_cost = st.columns([2, 1])
    with col_slider:
        max_li = st.slider("Max Posts to Fetch", 10, 999, 50, step=10, key="max_li")
    with col_cost:
        est_cost = max_li * COST_PER_LINKEDIN_POST
        st.metric("💰 Est. Cost", f"${est_cost:.2f}")

    st.markdown("---")
    if st.button("🚀 Scrape LinkedIn", use_container_width=True, key="btn_li"):
        if not keyword_li.strip():
            st.error("Please enter a keyword.")
        elif not api_token.strip():
            st.error("Please provide your Apify API Token in the sidebar.")
        else:
            st.session_state.accumulate_mode = accumulate_mode
            _run_scrape("linkedin", keyword_li, period_li, dates_li, api_token, max_li, debug_mode)

    _render_results("linkedin")


# ━━ TAB 2: X ━━
with tab_x:
    col_kw, col_tf = st.columns([2.5, 1])
    with col_kw:
        keyword_x = st.text_input(
            "Keyword / #hashtag / @mention",
            value="Polaris School of Technology",
            key="kw_x",
        )
    with col_tf:
        period_x = st.selectbox(
            "Time Window",
            ["today", "past-24h", "past-week", "past-month", "custom"],
            key="pd_x",
        )

    dates_x = None
    if period_x == "custom":
        dates_x = st.date_input(
            "📅 Date Range (Start → End)",
            value=(datetime.today().date() - timedelta(days=7), datetime.today().date()),
            max_value=datetime.today().date(),
            key="dates_x",
        )
        if isinstance(dates_x, (list, tuple)):
            dates_x = list(dates_x)
        else:
            dates_x = [dates_x]

    col_slider, col_cost = st.columns([2, 1])
    with col_slider:
        max_x = st.slider("Max Tweets to Fetch", 10, 999, 50, step=10, key="max_x")
    with col_cost:
        est_cost_x = max_x * COST_PER_X_TWEET
        st.metric("💰 Est. Cost", f"${est_cost_x:.4f}")

    st.markdown("---")
    if st.button("🚀 Scrape X (Twitter)", use_container_width=True, key="btn_x"):
        if not keyword_x.strip():
            st.error("Please enter a keyword.")
        elif not api_token.strip():
            st.error("Please provide your Apify API Token in the sidebar.")
        else:
            st.session_state.accumulate_mode = accumulate_mode
            _run_scrape("x", keyword_x, period_x, dates_x, api_token, max_x, debug_mode)

    _render_results("x")


# ━━ TAB 3: HOW TO USE ━━
with tab_help:
    st.markdown("""
### 🎯 What This Tool Does
Tracks every LinkedIn and X (Twitter) post that mentions Polaris School of Technology. 
Lets you filter by date range, see who posted, and download a clean Excel with all columns your team needs.

---

### 🔑 Step 1 — Get Your Apify API Key
1. Go to [apify.com](https://apify.com) → Sign up free
2. Settings → Integrations → **API Token**
3. Paste it in the sidebar

---

### 📅 Step 2 — Set Date Range
- **today** — posts from midnight IST to now
- **past-24h / past-week / past-month** — standard windows
- **custom** — pick any start and end date using the calendar picker

---

### 🚀 Step 3 — Scrape
Click the blue button. Wait ~1–3 minutes while Apify runs the actor.

---

### 📥 Step 4 — Download Excel
The Excel file has **4 tabs**:
| Tab | Content |
|-----|---------|
| All Posts | Every post scraped |
| Tags Polaris | Only posts where "Polaris" appears in the text |
| LinkedIn | LinkedIn posts only |
| X (Twitter) | X posts only |

**Columns in every tab:**
`Date · Time · Account Name · Profile URL · Post Link · Post Text (Preview) · Reactions · Comments · Reposts · Est. Impressions · Tags Polaris · Platform · Scraped At`

---

### ☁️ Best Way to Host This
**Streamlit Community Cloud** — 100% free, no server needed:
1. Push this folder to a GitHub repo (public or private)
2. Go to [share.streamlit.io](https://share.streamlit.io) → New App
3. Point to your repo and `app.py`
4. Add `APIFY_TOKEN` in Secrets (Settings → Secrets)
5. Your app gets a permanent URL like `https://yourname-polaris-tracker.streamlit.app`

""")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# COMBINED EXPORT — below tabs
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
_render_combined_download()
